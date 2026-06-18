#!/usr/bin/env python3
"""
Generate Security Report — Excel Workbook
Produces: Vulnerability Test Results/findings.xlsx
          Vulnerability Test Results/endpoint-inventory.xlsx

Requirements: pip install openpyxl
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

OUTPUT_DIR = "Vulnerability Test Results"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ──────────────────────────────────────────────
# Color palette
# ──────────────────────────────────────────────
CRITICAL_COLOR = "C62828"
HIGH_COLOR = "E65100"
MEDIUM_COLOR = "F9A825"
LOW_COLOR = "558B2F"
INFO_COLOR = "1565C0"
HEADER_COLOR = "1A237E"
ALT_ROW_COLOR = "E8EAF6"
WHITE = "FFFFFF"

def header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def severity_fill(severity: str):
    mapping = {
        "CRITICAL": CRITICAL_COLOR,
        "HIGH": HIGH_COLOR,
        "MEDIUM": MEDIUM_COLOR,
        "LOW": LOW_COLOR,
    }
    return PatternFill("solid", fgColor=mapping.get(severity.upper(), "CCCCCC"))

def thin_border():
    side = Side(style="thin", color="B0BEC5")
    return Border(left=side, right=side, top=side, bottom=side)

def bold_white(size=11):
    return Font(name="Calibri", bold=True, color=WHITE, size=size)

def normal_font(size=10, bold=False, color="000000"):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def center_align():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_align():
    return Alignment(horizontal="left", vertical="top", wrap_text=True)

def set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


# ──────────────────────────────────────────────
# Data
# ──────────────────────────────────────────────
FINDINGS = [
    {
        "id": "FINDING-001",
        "severity": "CRITICAL",
        "type": "Missing Authentication",
        "file": "All Controllers",
        "endpoint": "All API endpoints",
        "description": "No Spring Security filter chain configured. Every API endpoint is publicly accessible without any credentials.",
        "exploitation": "curl http://localhost:8080/user — returns all users. curl -X DELETE http://localhost:8080/user/delete/1 — deletes any user.",
        "impact": "Full account takeover, data breach of all PII, deletion of any record.",
        "fix": "Add spring-boot-starter-security + JWT. Create SecurityFilterChain protecting all routes. Issue JWT on login.",
        "cvss": "10.0",
    },
    {
        "id": "FINDING-002",
        "severity": "CRITICAL",
        "type": "Dangerous CORS (Wildcard)",
        "file": "All Controllers",
        "endpoint": "All API endpoints",
        "description": "@CrossOrigin(\"*\") on every controller allows any website to make cross-origin requests.",
        "exploitation": "Malicious page fetches http://api/user from any origin, exfiltrating all user data.",
        "impact": "Cross-site data exfiltration, CSRF attacks, full API access from any domain.",
        "fix": "Replace @CrossOrigin(\"*\") with specific origins. Configure CorsConfigurationSource globally.",
        "cvss": "9.1",
    },
    {
        "id": "FINDING-003",
        "severity": "CRITICAL",
        "type": "IDOR (Insecure Direct Object Reference)",
        "file": "All Controllers",
        "endpoint": "All endpoints with {id} path variables",
        "description": "No ownership verification on any resource endpoint. Any caller can read/modify/delete any other user's data by changing the ID.",
        "exploitation": "curl http://localhost:8080/user/1 — any user reads any other user's profile. DELETE /user/delete/5 from any session.",
        "impact": "Full horizontal and vertical privilege escalation. All user data accessible to anonymous requests.",
        "fix": "Extract authenticated user ID from JWT. Verify resource ownership before each operation. Return 403 on mismatch.",
        "cvss": "9.8",
    },
    {
        "id": "FINDING-004",
        "severity": "HIGH",
        "type": "Hardcoded Credentials",
        "file": "server/src/main/resources/application.properties",
        "endpoint": "N/A",
        "description": "MySQL root account with empty password hardcoded and committed to repository.",
        "exploitation": "Anyone with repo access connects directly: mysql -u root -h server-ip smartparking",
        "impact": "Full database compromise, potential OS escalation via MySQL FILE privilege.",
        "fix": "Use environment variables for DB credentials. Create least-privilege DB user. Add to .gitignore.",
        "cvss": "8.8",
    },
    {
        "id": "FINDING-005",
        "severity": "HIGH",
        "type": "Unsafe File Upload / Path Traversal",
        "file": "ParkingPlaceController.java, VehicleController.java",
        "endpoint": "POST /parking/{lenderId}, POST /vehicle/{userId}",
        "description": "Raw original filename from client used to construct server file path without sanitization. Any file type accepted.",
        "exploitation": "Upload file named ../../application.properties to overwrite config. Upload shell.jsp for RCE.",
        "impact": "Remote Code Execution, file system compromise, configuration overwrite.",
        "fix": "Use UUID filenames. Validate MIME type and extension. Store outside web root. Apply UserService.saveImage() pattern.",
        "cvss": "8.6",
    },
    {
        "id": "FINDING-006",
        "severity": "HIGH",
        "type": "SQL/Plaintext Password Comparison Risk",
        "file": "UserService.java, LenderService.java",
        "endpoint": "N/A (service layer)",
        "description": "findByEmailAndPassword() methods exist that compare plaintext passwords in DB queries, bypassing BCrypt.",
        "exploitation": "If these code paths are ever called, plaintext password is sent directly to DB comparison, enabling SQL injection.",
        "impact": "Authentication bypass, SQL injection risk, security model breakdown.",
        "fix": "Remove findByEmailAndPassword methods. Always authenticate via findByEmail() + passwordEncoder.matches().",
        "cvss": "7.5",
    },
    {
        "id": "FINDING-007",
        "severity": "HIGH",
        "type": "Password/PII Exposure in Responses",
        "file": "UserController.java, LenderController.java",
        "endpoint": "GET /user, GET /user/{id}, GET /lender, GET /lender/{id}",
        "description": "GET endpoints return full entity including BCrypt-hashed password, email, mobile, and address of all users.",
        "exploitation": "curl http://localhost:8080/user — returns all hashed passwords. Run hashcat -m 3200 to crack them offline.",
        "impact": "PII breach, offline password cracking, GDPR/PCI-DSS violations.",
        "fix": "Add @JsonProperty(access=WRITE_ONLY) to password field. Use DTO projections for list responses.",
        "cvss": "7.5",
    },
    {
        "id": "FINDING-008",
        "severity": "HIGH",
        "type": "No Rate Limiting / Brute-Force Protection",
        "file": "AuthenticationController.java",
        "endpoint": "POST /login/user, POST /login/lender",
        "description": "Login endpoints have no rate limiting, account lockout, or CAPTCHA. Unlimited password attempts allowed.",
        "exploitation": "Python script sends 10000 login requests with password wordlist. No throttling applied.",
        "impact": "Account brute-force, credential stuffing attacks, account takeover.",
        "fix": "Implement Bucket4j rate limiting. Add exponential backoff. Lock accounts after N failures.",
        "cvss": "7.3",
    },
    {
        "id": "FINDING-009",
        "severity": "MEDIUM",
        "type": "SQL Query Logging in Production",
        "file": "server/src/main/resources/application.properties",
        "endpoint": "N/A",
        "description": "spring.jpa.show-sql=true logs all SQL queries including parameters to stdout/production logs.",
        "exploitation": "Log access reveals user emails, query structure, and application internals.",
        "impact": "PII leakage in logs, security model exposure.",
        "fix": "Set spring.jpa.show-sql=false in production. Use separate dev profile for SQL debugging.",
        "cvss": "5.3",
    },
    {
        "id": "FINDING-010",
        "severity": "MEDIUM",
        "type": "Client-Controlled Booking Status",
        "file": "BookingController.java",
        "endpoint": "POST /booking/update/status/{bookingId}",
        "description": "Booking status accepted as free-text string from client. Any value accepted with no validation or workflow enforcement.",
        "exploitation": "curl -X POST '/booking/update/status/5?status=Confirmed' — skips payment approval flow.",
        "impact": "Financial fraud, booking manipulation, workflow bypass.",
        "fix": "Define BookingStatus enum. Validate transitions server-side. Add authorization checks.",
        "cvss": "6.5",
    },
    {
        "id": "FINDING-011",
        "severity": "MEDIUM",
        "type": "Vulnerable Dependency (Apache Velocity 1.7)",
        "file": "server/pom.xml",
        "endpoint": "N/A",
        "description": "Apache Velocity 1.7 (2010, EOL) with CVE-2020-13936 (CVSS 9.8) — Server-Side Template Injection allowing RCE.",
        "exploitation": "If user input reaches a Velocity template, attacker escapes sandbox and executes arbitrary Java.",
        "impact": "Remote Code Execution on the server.",
        "fix": "Replace with velocity-engine-core:2.3+ or migrate to Thymeleaf.",
        "cvss": "9.8 (CVE)",
    },
    {
        "id": "FINDING-012",
        "severity": "MEDIUM",
        "type": "Missing Security Headers",
        "file": "Application-wide (no SecurityFilterChain)",
        "endpoint": "All responses",
        "description": "No X-Content-Type-Options, X-Frame-Options, CSP, HSTS, or Referrer-Policy headers present.",
        "exploitation": "Enables clickjacking, MIME sniffing, and cross-site script injection.",
        "impact": "Clickjacking, XSS amplification, data leakage.",
        "fix": "Add Spring Security. Configure security headers via http.headers() configuration.",
        "cvss": "4.3",
    },
    {
        "id": "FINDING-013",
        "severity": "MEDIUM",
        "type": "Excessive File Upload Size (DoS)",
        "file": "server/src/main/resources/application.properties",
        "endpoint": "POST /parking/{lenderId}, POST /vehicle/{userId}",
        "description": "max-file-size=100MB allows anonymous DoS attacks by uploading large files repeatedly.",
        "exploitation": "Loop uploading 100MB files fills disk, crashes server out of memory.",
        "impact": "Denial of Service, disk exhaustion.",
        "fix": "Reduce to max-file-size=5MB. Add authentication to upload endpoints.",
        "cvss": "5.0",
    },
    {
        "id": "FINDING-014",
        "severity": "LOW",
        "type": "Stack Trace Logging (e.printStackTrace())",
        "file": "AuthenticationController.java, BookingController.java",
        "endpoint": "Multiple",
        "description": "e.printStackTrace() dumps full stack traces to stdout, revealing implementation details.",
        "exploitation": "Log access reveals class names, library versions, DB query structure.",
        "impact": "Information leakage, aids targeted attacks.",
        "fix": "Use SLF4J Logger with log.error(\"message\", e) instead of e.printStackTrace().",
        "cvss": "3.1",
    },
    {
        "id": "FINDING-015",
        "severity": "LOW",
        "type": "Missing Input Validation on User Signup",
        "file": "UserController.java",
        "endpoint": "POST /user/signup",
        "description": "No @Valid annotation and no validation constraints on the User entity — accepts empty names, invalid emails, weak passwords.",
        "exploitation": "Register with email='x', password='1', name='' — bypasses any front-end validation.",
        "impact": "Data integrity issues, security control bypass.",
        "fix": "Add @Valid to @RequestBody. Add @NotBlank, @Email, @Size(min=8) to User entity or registration DTO.",
        "cvss": "4.0",
    },
    {
        "id": "FINDING-016",
        "severity": "LOW",
        "type": "No Pagination on List Endpoints",
        "file": "UserController.java, LenderController.java, BookingController.java",
        "endpoint": "GET /user, GET /lender, GET /booking/all, GET /rating/all",
        "description": "List endpoints return all records with no pagination, causing memory exhaustion and full data exposure.",
        "exploitation": "curl /user returns entire user database in one response.",
        "impact": "DoS via memory exhaustion, excessive data exposure.",
        "fix": "Add Pageable parameter to all list endpoints and JPA repository methods.",
        "cvss": "3.7",
    },
    {
        "id": "FINDING-017",
        "severity": "LOW",
        "type": "Insecure Client-Side Storage (localStorage)",
        "file": "client/src (login components)",
        "endpoint": "N/A (frontend)",
        "description": "Full user object including PII stored in localStorage, accessible to any JavaScript (XSS risk).",
        "exploitation": "If XSS exists: document.localStorage.getItem('user') exfiltrates all stored PII.",
        "impact": "Session hijacking, PII exposure via XSS.",
        "fix": "Store only JWT token. Use httpOnly cookies for sensitive tokens. Never store PII in localStorage.",
        "cvss": "4.3",
    },
]

ENDPOINTS = [
    ("POST", "/login/user", "No", "Public", "AuthenticationController.java"),
    ("POST", "/login/lender", "No", "Public", "AuthenticationController.java"),
    ("POST", "/user/signup", "No", "Public", "UserController.java"),
    ("GET", "/user", "No (should be Yes)", "Admin", "UserController.java"),
    ("GET", "/user/{id}", "No (should be Yes)", "Admin/Self", "UserController.java"),
    ("PUT", "/user/{userId}", "No (should be Yes)", "Self", "UserController.java"),
    ("DELETE", "/user/delete/{id}", "No (should be Yes)", "Admin", "UserController.java"),
    ("POST", "/lender/signup", "No", "Public", "LenderController.java"),
    ("GET", "/lender", "No (should be Yes)", "Admin", "LenderController.java"),
    ("GET", "/lender/{id}", "No (should be Yes)", "Admin/Self", "LenderController.java"),
    ("PUT", "/lender/update/{lenderId}", "No (should be Yes)", "Self", "LenderController.java"),
    ("DELETE", "/lender/delete/{id}", "No (should be Yes)", "Admin", "LenderController.java"),
    ("POST", "/parking/{lenderId}", "No (should be Yes)", "Lender", "ParkingPlaceController.java"),
    ("GET", "/parking", "No", "Public", "ParkingPlaceController.java"),
    ("GET", "/parking/place/{lenderId}", "No", "Public", "ParkingPlaceController.java"),
    ("GET", "/parking/area/{areaName}", "No", "Public", "ParkingPlaceController.java"),
    ("DELETE", "/parking/delete/{id}", "No (should be Yes)", "Admin/Lender", "ParkingPlaceController.java"),
    ("POST", "/booking/parking/{lenderId}/book", "No (should be Yes)", "User", "BookingController.java"),
    ("POST", "/booking/release/{bookingId}", "No (should be Yes)", "Self/Admin", "BookingController.java"),
    ("POST", "/booking/update/status/{bookingId}", "No (should be Yes)", "Admin", "BookingController.java"),
    ("GET", "/booking/user/{userId}", "No (should be Yes)", "Self", "BookingController.java"),
    ("GET", "/booking/all", "No (should be Yes)", "Admin", "BookingController.java"),
    ("DELETE", "/booking/{id}", "No (should be Yes)", "Admin", "BookingController.java"),
    ("GET", "/booking/lender/{lenderId}", "No (should be Yes)", "Lender/Admin", "BookingController.java"),
    ("GET", "/booking/{id}", "No (should be Yes)", "Self/Admin", "BookingController.java"),
    ("POST", "/vehicle/{userId}", "No (should be Yes)", "Self", "VehicleController.java"),
    ("GET", "/vehicle/{userId}", "No (should be Yes)", "Self", "VehicleController.java"),
    ("DELETE", "/vehicle/delete/{id}", "No (should be Yes)", "Self", "VehicleController.java"),
    ("POST", "/rating/add", "No (should be Yes)", "User", "RatingController.java"),
    ("GET", "/rating/all", "No", "Public", "RatingController.java"),
    ("GET", "/rating/parking/{lenderId}", "No", "Public", "RatingController.java"),
    ("GET", "/rating/user/{userId}", "No", "Public", "RatingController.java"),
    ("DELETE", "/rating/{id}", "No (should be Yes)", "Admin/Self", "RatingController.java"),
]

DEPENDENCIES = [
    ("Apache Velocity", "org.apache.velocity:velocity", "1.7", "CRITICAL", "CVE-2020-13936 (CVSS 9.8) — SSTI/RCE", "Replace with velocity-engine-core:2.3+"),
    ("Spring Boot", "spring-boot-starter-parent", "3.3.4", "LOW", "None known", "Monitor for 3.4.x"),
    ("Spring Security Crypto", "spring-security-crypto", "6.3.3", "LOW", "None", "Current"),
    ("MySQL Connector", "mysql-connector-j", "managed", "LOW", "None known", "Verify latest"),
    ("Lombok", "lombok", "1.18.38", "LOW", "None", "Current"),
    ("React", "react", "18.3.1", "LOW", "None", "Current"),
    ("React Router DOM", "react-router-dom", "6.26.2", "LOW", "None", "Current"),
    ("Axios", "axios", "1.7.7 / 1.18.0", "LOW", "None", "Current"),
    ("Vite", "vite", "5.4.8", "MEDIUM", "None critical; Vite 6.x available", "Upgrade to 6.x"),
    ("React (mobile)", "react", "19.2.3", "MEDIUM", "React 19 still stabilizing", "Monitor for stable release"),
    ("TailwindCSS", "tailwindcss", "3.4.13", "LOW", "None", "Current"),
]


# ──────────────────────────────────────────────
# Workbook 1: findings.xlsx
# ──────────────────────────────────────────────
def create_findings_workbook():
    wb = Workbook()

    # ── Sheet 1: Security Findings ──
    ws1 = wb.active
    ws1.title = "Security Findings"

    headers = ["ID", "Severity", "Vulnerability Type", "File Path", "Endpoint",
               "Description", "Exploitation Scenario", "Impact", "Recommended Fix", "CVSS"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill(HEADER_COLOR)
        cell.font = bold_white(11)
        cell.alignment = center_align()
        cell.border = thin_border()

    for row_idx, f in enumerate(FINDINGS, 2):
        fill = header_fill(ALT_ROW_COLOR) if row_idx % 2 == 0 else header_fill(WHITE)
        values = [f["id"], f["severity"], f["type"], f["file"], f["endpoint"],
                  f["description"], f["exploitation"], f["impact"], f["fix"], f["cvss"]]
        for col_idx, val in enumerate(values, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = left_align()
            cell.border = thin_border()
            if col_idx == 2:  # Severity column
                cell.fill = severity_fill(f["severity"])
                cell.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
                cell.alignment = center_align()
            else:
                cell.fill = fill
                cell.font = normal_font()

    set_col_widths(ws1, {
        "A": 16, "B": 12, "C": 28, "D": 35, "E": 35,
        "F": 55, "G": 50, "H": 45, "I": 55, "J": 12,
    })
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = ws1.dimensions
    ws1.row_dimensions[1].height = 30

    # ── Sheet 2: Endpoint Inventory ──
    ws2 = wb.create_sheet("Endpoint Inventory")
    ep_headers = ["HTTP Method", "Endpoint", "Auth Required", "Expected Roles", "Controller File"]
    for col, h in enumerate(ep_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill(HEADER_COLOR)
        cell.font = bold_white()
        cell.alignment = center_align()
        cell.border = thin_border()

    method_colors = {"GET": "1976D2", "POST": "388E3C", "PUT": "F57C00", "DELETE": "C62828"}
    for row_idx, (method, endpoint, auth, roles, ctrl) in enumerate(ENDPOINTS, 2):
        bg = ALT_ROW_COLOR if row_idx % 2 == 0 else WHITE
        for col_idx, val in enumerate([method, endpoint, auth, roles, ctrl], 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border()
            cell.alignment = left_align()
            if col_idx == 1:
                cell.fill = header_fill(method_colors.get(method, "607D8B"))
                cell.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
                cell.alignment = center_align()
            elif col_idx == 3 and "should be" in val:
                cell.fill = severity_fill("HIGH")
                cell.font = Font(name="Calibri", bold=True, color=WHITE, size=9)
            else:
                cell.fill = header_fill(bg)
                cell.font = normal_font()

    set_col_widths(ws2, {"A": 14, "B": 45, "C": 25, "D": 22, "E": 38})
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions

    # ── Sheet 3: Dependency Vulnerabilities ──
    ws3 = wb.create_sheet("Dependency Vulnerabilities")
    dep_headers = ["Package Name", "Artifact ID", "Version", "Risk", "CVEs / Notes", "Recommended Action"]
    for col, h in enumerate(dep_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.fill = header_fill(HEADER_COLOR)
        cell.font = bold_white()
        cell.alignment = center_align()
        cell.border = thin_border()

    for row_idx, dep in enumerate(DEPENDENCIES, 2):
        bg = ALT_ROW_COLOR if row_idx % 2 == 0 else WHITE
        for col_idx, val in enumerate(dep, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border()
            cell.alignment = left_align()
            if col_idx == 4:
                cell.fill = severity_fill(dep[3])
                cell.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
                cell.alignment = center_align()
            else:
                cell.fill = header_fill(bg)
                cell.font = normal_font()

    set_col_widths(ws3, {"A": 22, "B": 35, "C": 15, "D": 12, "E": 45, "F": 38})
    ws3.freeze_panes = "A2"

    # ── Sheet 4: Risk Summary ──
    ws4 = wb.create_sheet("Risk Summary")
    ws4["A1"] = "Smart Parking & Reservation System — Security Risk Summary"
    ws4["A1"].font = Font(name="Calibri", bold=True, size=16, color=HEADER_COLOR)
    ws4["A1"].alignment = center_align()
    ws4.merge_cells("A1:D1")
    ws4.row_dimensions[1].height = 40

    ws4["A3"] = f"Assessment Date: {datetime.now().strftime('%Y-%m-%d')}"
    ws4["A3"].font = normal_font(size=12)
    ws4["A4"] = "Repository: https://github.com/Viru-6281/pdd-main"
    ws4["A4"].font = normal_font(size=12)
    ws4["A5"] = "Overall Security Score: 8 / 100"
    ws4["A5"].font = Font(name="Calibri", bold=True, size=14, color=CRITICAL_COLOR)

    summary_data = [
        ("", "", "", ""),
        ("Severity", "Count", "% of Total", "Status"),
        ("CRITICAL", 3, "17.6%", "🔴 Requires immediate remediation"),
        ("HIGH", 5, "29.4%", "🟠 Requires urgent action"),
        ("MEDIUM", 5, "29.4%", "🟡 Address in next sprint"),
        ("LOW", 4, "23.5%", "🟢 Routine improvement"),
        ("TOTAL", 17, "100%", "NOT PRODUCTION READY"),
    ]

    start_row = 7
    for r_idx, row in enumerate(summary_data, start_row):
        for c_idx, val in enumerate(row, 1):
            cell = ws4.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border()
            cell.alignment = center_align()
            if r_idx == start_row + 1:  # Header row
                cell.fill = header_fill(HEADER_COLOR)
                cell.font = bold_white()
            elif r_idx == start_row + 2:
                cell.fill = severity_fill("CRITICAL")
                cell.font = Font(name="Calibri", bold=True, color=WHITE)
            elif r_idx == start_row + 3:
                cell.fill = severity_fill("HIGH")
                cell.font = Font(name="Calibri", bold=True, color=WHITE)
            elif r_idx == start_row + 4:
                cell.fill = severity_fill("MEDIUM")
                cell.font = Font(name="Calibri", bold=True, color="000000")
            elif r_idx == start_row + 5:
                cell.fill = severity_fill("LOW")
                cell.font = Font(name="Calibri", bold=True, color=WHITE)
            elif r_idx == start_row + 6:
                cell.fill = severity_fill("CRITICAL")
                cell.font = Font(name="Calibri", bold=True, color=WHITE)
            else:
                cell.font = normal_font()

    set_col_widths(ws4, {"A": 18, "B": 12, "C": 15, "D": 45})

    path = os.path.join(OUTPUT_DIR, "findings.xlsx")
    wb.save(path)
    print(f"✅ Saved: {path}")


# ──────────────────────────────────────────────
# Workbook 2: endpoint-inventory.xlsx (standalone)
# ──────────────────────────────────────────────
def create_endpoint_inventory():
    wb = Workbook()
    ws = wb.active
    ws.title = "API Endpoint Inventory"

    headers = ["#", "HTTP Method", "Endpoint Path", "Auth Required", "Expected Roles",
               "Controller File", "Security Risk"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill(HEADER_COLOR)
        cell.font = bold_white()
        cell.alignment = center_align()
        cell.border = thin_border()

    method_colors = {"GET": "1976D2", "POST": "388E3C", "PUT": "F57C00", "DELETE": "C62828"}
    for row_idx, (method, endpoint, auth, roles, ctrl) in enumerate(ENDPOINTS, 2):
        is_risky = "should be" in auth.lower()
        bg = "FFEBEE" if is_risky else (ALT_ROW_COLOR if row_idx % 2 == 0 else WHITE)
        risk = "⚠️ Unprotected" if is_risky else "✅ Public"

        for col_idx, val in enumerate([row_idx - 1, method, endpoint, auth, roles, ctrl, risk], 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border()
            cell.alignment = left_align()
            if col_idx == 2:
                cell.fill = header_fill(method_colors.get(method, "607D8B"))
                cell.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
                cell.alignment = center_align()
            elif col_idx == 7 and is_risky:
                cell.fill = severity_fill("HIGH")
                cell.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
                cell.alignment = center_align()
            else:
                cell.fill = header_fill(bg)
                cell.font = normal_font()

    set_col_widths(ws, {"A": 6, "B": 14, "C": 48, "D": 26, "E": 22, "F": 38, "G": 20})
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    path = os.path.join(OUTPUT_DIR, "endpoint-inventory.xlsx")
    wb.save(path)
    print(f"✅ Saved: {path}")


if __name__ == "__main__":
    print("🔐 Generating security reports...")
    create_findings_workbook()
    create_endpoint_inventory()
    print("\n✅ All reports generated in:", OUTPUT_DIR)
    print("  📄 findings.xlsx — Security findings + endpoint inventory + dependency + risk summary")
    print("  📄 endpoint-inventory.xlsx — Standalone API inventory")
