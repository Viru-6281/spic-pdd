#!/usr/bin/env python3
"""
Security Report Excel Generator
Generates endpoint-inventory.xlsx and findings.xlsx for Smart Parking & Reservation
Usage: python generate_reports.py
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.series import DataPoint
from datetime import datetime
import os

# ─────────────────────────────────────────────
# COLOR PALETTE
# ─────────────────────────────────────────────
COLORS = {
    "critical": "C0392B",
    "high":     "E67E22",
    "medium":   "F1C40F",
    "low":      "27AE60",
    "header":   "1A1A2E",
    "subheader":"16213E",
    "white":    "FFFFFF",
    "light_gray":"F5F5F5",
    "dark_text": "2C3E50",
    "border":   "BDC3C7",
    "pass":     "27AE60",
    "fail":     "C0392B",
    "info":     "2980B9",
}

def make_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def make_border(color="BDC3C7"):
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)

def header_font(size=11, bold=True, color="FFFFFF"):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def body_font(size=10, bold=False, color="2C3E50"):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def apply_header_row(ws, row, values, fill_color="1A1A2E"):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = make_fill(fill_color)
        cell.font = header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = make_border()

def apply_data_row(ws, row, values, fill_color="FFFFFF", bold=False):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = make_fill(fill_color)
        cell.font = body_font(bold=bold)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = make_border()

def set_col_widths(ws, widths):
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

# ─────────────────────────────────────────────
# DATA
# ─────────────────────────────────────────────
FINDINGS = [
    {
        "id": "F-001",
        "severity": "Critical",
        "type": "Wildcard CORS",
        "file": "controller/*.java",
        "endpoint": "All endpoints",
        "description": "All 7 controllers use @CrossOrigin(\"*\"), allowing any origin to read API responses.",
        "exploitation": "Malicious page at evil.com reads user data via AJAX from victim's session.",
        "impact": "Full data exfiltration of PII across origins.",
        "fix": "Replace @CrossOrigin(\"*\") with allowlist. Use global WebMvcConfigurer with specific origins.",
        "cwe": "CWE-346",
        "owasp": "A05:2021 - Security Misconfiguration"
    },
    {
        "id": "F-002",
        "severity": "Critical",
        "type": "Broken Authentication — No Server-Side Auth Enforcement",
        "file": "controller/AuthenticationController.java",
        "endpoint": "All /login/* and protected endpoints",
        "description": "Login returns user object; no JWT or session token enforced on subsequent requests. Any API call works without credentials.",
        "exploitation": "Call GET /booking/all or GET /user directly with no headers. Full data returned.",
        "impact": "Complete authentication bypass. All data publicly readable.",
        "fix": "Implement Spring Security + JWT. Protect all non-public endpoints with SecurityFilterChain.",
        "cwe": "CWE-306",
        "owasp": "A01:2021 - Broken Access Control"
    },
    {
        "id": "F-003",
        "severity": "Critical",
        "type": "IDOR — Insecure Direct Object Reference",
        "file": "controller/BookingController.java, UserController.java",
        "endpoint": "GET /booking/{id}, DELETE /booking/{id}, GET /user/{id}",
        "description": "Numeric resource IDs accepted with no ownership verification. Any caller can access any resource.",
        "exploitation": "Enumerate IDs from 1–1000 to harvest all user profiles and bookings.",
        "impact": "Horizontal privilege escalation. Mass PII breach.",
        "fix": "Extract authenticated user from JWT; verify resource ownership before returning data.",
        "cwe": "CWE-639",
        "owasp": "A01:2021 - Broken Access Control"
    },
    {
        "id": "F-004",
        "severity": "High",
        "type": "Path Traversal via File Upload",
        "file": "controller/ParkingPlaceController.java:56, VehicleController.java:47",
        "endpoint": "POST /parking/{lenderId}, POST /vehicle/{userId}",
        "description": "image.getOriginalFilename() used directly in Paths.get() without sanitisation.",
        "exploitation": "Upload file named ../../application.properties to overwrite server config.",
        "impact": "Remote config overwrite, potential RCE via file placement.",
        "fix": "Sanitise filename: UUID rename, extension whitelist, normalize and check path prefix.",
        "cwe": "CWE-22",
        "owasp": "A03:2021 - Injection"
    },
    {
        "id": "F-005",
        "severity": "High",
        "type": "Excessive Data Exposure",
        "file": "controller/BookingController.java:208",
        "endpoint": "GET /booking/all",
        "description": "Returns all bookings for all users in the system with no authentication or pagination.",
        "exploitation": "Automated scraping of endpoint dumps all bookings with user PII.",
        "impact": "Mass PII breach. GDPR violation.",
        "fix": "Restrict to ADMIN role. Implement pagination. Require authentication.",
        "cwe": "CWE-200",
        "owasp": "A02:2021 - Cryptographic Failures"
    },
    {
        "id": "F-006",
        "severity": "High",
        "type": "Hardcoded Empty Database Password",
        "file": "server/src/main/resources/application.properties:6",
        "endpoint": "N/A",
        "description": "spring.datasource.password= is empty and committed to the repository.",
        "exploitation": "If DB port is exposed, connect without credentials. If repo is public, attacker knows.",
        "impact": "Full database compromise.",
        "fix": "Use environment variable: spring.datasource.password=${DB_PASSWORD}",
        "cwe": "CWE-259",
        "owasp": "A02:2021 - Cryptographic Failures"
    },
    {
        "id": "F-007",
        "severity": "High",
        "type": "Information Disclosure in Error Messages",
        "file": "controller/LenderController.java:120, VehicleController.java:87",
        "endpoint": "GET /lender/{id}, GET /vehicle/{userId}",
        "description": "User-controlled path variables echoed in error messages: 'not found for id' + id",
        "exploitation": "Confirm valid IDs by response variation. Enumerate all valid entity IDs.",
        "impact": "Information disclosure aids IDOR exploitation.",
        "fix": "Use generic error messages. Never echo user input in API responses.",
        "cwe": "CWE-209",
        "owasp": "A05:2021 - Security Misconfiguration"
    },
    {
        "id": "F-008",
        "severity": "High",
        "type": "Missing Rate Limiting on Login Endpoints",
        "file": "controller/AuthenticationController.java",
        "endpoint": "POST /login/user, POST /login/lender",
        "description": "No rate limiting, account lockout, or CAPTCHA on login endpoints.",
        "exploitation": "Submit 100,000 password attempts against known email without throttling.",
        "impact": "Account takeover via brute force.",
        "fix": "Use Bucket4j or nginx rate limiting. Implement account lockout after 5 failed attempts.",
        "cwe": "CWE-307",
        "owasp": "A07:2021 - Identification and Authentication Failures"
    },
    {
        "id": "F-009",
        "severity": "Medium",
        "type": "Missing Security Headers",
        "file": "server/src/main/resources/application.properties",
        "endpoint": "All endpoints",
        "description": "X-Content-Type-Options, X-Frame-Options, HSTS, CSP headers missing from all responses.",
        "exploitation": "Clickjacking via iframe embedding. MIME sniffing attacks.",
        "impact": "Clickjacking, content sniffing, downgrade attacks.",
        "fix": "Configure Spring Security headers: frame-options=DENY, content-type-options, HSTS.",
        "cwe": "CWE-693",
        "owasp": "A05:2021 - Security Misconfiguration"
    },
    {
        "id": "F-010",
        "severity": "Medium",
        "type": "Debug SQL Logging Enabled",
        "file": "server/src/main/resources/application.properties:11",
        "endpoint": "N/A",
        "description": "spring.jpa.show-sql=true prints all SQL queries to stdout including schema details.",
        "exploitation": "Read CI/CD logs to extract schema, table names, query patterns.",
        "impact": "Schema information leakage.",
        "fix": "Set spring.jpa.show-sql=false in production profile.",
        "cwe": "CWE-532",
        "owasp": "A09:2021 - Security Logging Failures"
    },
    {
        "id": "F-011",
        "severity": "Medium",
        "type": "Hardcoded localhost API URL",
        "file": "client/src/AxiosInstance.jsx:5",
        "endpoint": "All frontend API calls",
        "description": "baseURL: 'http://localhost:8080/' — production deployment calls non-existent server over HTTP.",
        "exploitation": "All API calls in production fail silently. Credentials sent over HTTP (not HTTPS).",
        "impact": "Non-functional production deployment. Credential exposure over HTTP.",
        "fix": "Use VITE_API_BASE_URL environment variable. Enforce HTTPS.",
        "cwe": "CWE-319",
        "owasp": "A02:2021 - Cryptographic Failures"
    },
    {
        "id": "F-012",
        "severity": "Medium",
        "type": "BrowserRouter — GitHub Pages 404 on Refresh",
        "file": "client/src/App.jsx:1,28",
        "endpoint": "All frontend routes",
        "description": "BrowserRouter relies on HTML5 History API. GitHub Pages returns 404 on direct URL access.",
        "exploitation": "All non-root routes return 404. Users cannot share or bookmark links.",
        "impact": "Broken deployment. All routes except / are inaccessible.",
        "fix": "Replace BrowserRouter with HashRouter from react-router-dom.",
        "cwe": "CWE-601",
        "owasp": "A05:2021 - Security Misconfiguration"
    },
    {
        "id": "F-013",
        "severity": "Medium",
        "type": "Oversized File Upload Limit (100MB)",
        "file": "server/src/main/resources/application.properties:16-18",
        "endpoint": "POST /parking/{lenderId}, POST /vehicle/{userId}",
        "description": "100MB upload limit with no content-type or magic-byte validation.",
        "exploitation": "Repeated 100MB uploads exhaust disk space. Upload executable disguised as image.",
        "impact": "Denial of Service via disk exhaustion.",
        "fix": "Reduce limit to 5MB. Validate MIME type and magic bytes server-side.",
        "cwe": "CWE-400",
        "owasp": "A04:2021 - Insecure Design"
    },
    {
        "id": "F-014",
        "severity": "Low",
        "type": "Stack Trace Disclosure",
        "file": "All controllers — all catch blocks",
        "endpoint": "N/A",
        "description": "e.printStackTrace() in all catch blocks leaks internal class names and stack frames to logs.",
        "exploitation": "Read server logs to identify framework versions and internal structure.",
        "impact": "Internal information leakage aids reconnaissance.",
        "fix": "Replace with SLF4J: log.error(\"message\", e). Filter log levels in production.",
        "cwe": "CWE-209",
        "owasp": "A09:2021 - Security Logging Failures"
    },
    {
        "id": "F-015",
        "severity": "Low",
        "type": "EOL Dependency — Apache Velocity 1.7",
        "file": "server/pom.xml:68-71",
        "endpoint": "N/A",
        "description": "Apache Velocity 1.7 (2010, EOL). CVE-2020-13936: template injection via ClassInfo.",
        "exploitation": "If Velocity renders user-controlled content, arbitrary Java method invocation.",
        "impact": "Template injection → potential RCE.",
        "fix": "Upgrade to velocity-engine-core:2.3 or migrate to Thymeleaf.",
        "cwe": "CWE-94",
        "owasp": "A06:2021 - Vulnerable Components"
    },
    {
        "id": "F-016",
        "severity": "Low",
        "type": "Missing Status Enum Validation",
        "file": "controller/BookingController.java:133",
        "endpoint": "POST /booking/update/status/{bookingId}",
        "description": "booking.setStatus(status) — status accepted without validation against allowed values.",
        "exploitation": "Set status to any arbitrary string to bypass business logic.",
        "impact": "Business logic bypass. Fake status values.",
        "fix": "Validate against Set.of(\"PENDING\",\"CONFIRMED\",\"CANCELLED\",\"COMPLETED\").",
        "cwe": "CWE-20",
        "owasp": "A03:2021 - Injection"
    },
]

ENDPOINTS = [
    # Authentication
    ("POST", "/login/user", "No", "None (Public)", "AuthenticationController.java", "User login with email+password"),
    ("POST", "/login/lender", "No", "None (Public)", "AuthenticationController.java", "Lender login with email+password"),
    # User
    ("POST", "/user/signup", "No", "None (Public)", "UserController.java", "User registration"),
    ("PUT", "/user/{userId}", "No ⚠️", "Owner", "UserController.java", "Update user profile with image"),
    ("GET", "/user", "No ⚠️", "Admin", "UserController.java", "Get all users — NO AUTH"),
    ("GET", "/user/{id}", "No ⚠️", "Owner/Admin", "UserController.java", "Get user by ID — NO AUTH"),
    ("DELETE", "/user/delete/{id}", "No ⚠️", "Admin", "UserController.java", "Delete user — NO AUTH"),
    # Lender
    ("POST", "/lender/signup", "No", "None (Public)", "LenderController.java", "Lender registration"),
    ("PUT", "/lender/update/{lenderId}", "No ⚠️", "Owner", "LenderController.java", "Update lender profile"),
    ("GET", "/lender", "No ⚠️", "Admin", "LenderController.java", "Get all lenders — NO AUTH"),
    ("GET", "/lender/{id}", "No ⚠️", "Owner/Admin", "LenderController.java", "Get lender by ID — NO AUTH"),
    ("DELETE", "/lender/delete/{id}", "No ⚠️", "Admin", "LenderController.java", "Delete lender — NO AUTH"),
    # Booking
    ("POST", "/booking/parking/{lenderId}/book", "No ⚠️", "User", "BookingController.java", "Book parking place"),
    ("POST", "/booking/release/{bookingId}", "No ⚠️", "User/Lender", "BookingController.java", "Release parking spot"),
    ("POST", "/booking/update/status/{bookingId}", "No ⚠️", "Lender/Admin", "BookingController.java", "Update booking status"),
    ("GET", "/booking/user/{userId}", "No ⚠️", "Owner", "BookingController.java", "Get bookings by user — NO AUTH"),
    ("GET", "/booking/all", "No ⚠️", "Admin", "BookingController.java", "⚠️ Get ALL bookings — NO AUTH"),
    ("GET", "/booking/{id}", "No ⚠️", "Owner/Admin", "BookingController.java", "Get booking by ID — NO AUTH"),
    ("DELETE", "/booking/{id}", "No ⚠️", "Owner/Admin", "BookingController.java", "Delete booking — NO AUTH"),
    ("GET", "/booking/lender/{lenderId}", "No ⚠️", "Lender", "BookingController.java", "Get bookings by lender — NO AUTH"),
    # Parking
    ("POST", "/parking/{lenderId}", "No ⚠️", "Lender", "ParkingPlaceController.java", "Add parking place with image"),
    ("GET", "/parking", "No", "None (Public)", "ParkingPlaceController.java", "Get all parking places"),
    ("GET", "/parking/place/{lenderId}", "No ⚠️", "Lender", "ParkingPlaceController.java", "Get parking places by lender"),
    ("GET", "/parking/area/{areaName}", "No", "None (Public)", "ParkingPlaceController.java", "Search by area name"),
    ("DELETE", "/parking/delete/{id}", "No ⚠️", "Lender/Admin", "ParkingPlaceController.java", "Delete parking place — NO AUTH"),
    # Vehicle
    ("POST", "/vehicle/{userId}", "No ⚠️", "User", "VehicleController.java", "Add vehicle with image upload"),
    ("GET", "/vehicle/{userId}", "No ⚠️", "Owner", "VehicleController.java", "Get vehicles by user — NO AUTH"),
    ("DELETE", "/vehicle/delete/{id}", "No ⚠️", "Owner", "VehicleController.java", "Delete vehicle — NO AUTH"),
    # Rating
    ("POST", "/rating/add", "No ⚠️", "User", "RatingController.java", "Add rating and comment"),
    ("GET", "/rating/all", "No ⚠️", "Admin", "RatingController.java", "Get all ratings — NO AUTH"),
    ("GET", "/rating/parking/{lenderId}", "No ⚠️", "Public", "RatingController.java", "Get ratings by lender"),
    ("GET", "/rating/user/{userId}", "No ⚠️", "Owner", "RatingController.java", "Get ratings by user — NO AUTH"),
    ("DELETE", "/rating/{id}", "No ⚠️", "Admin", "RatingController.java", "Delete rating — NO AUTH"),
]

DEPENDENCIES = [
    ("velocity", "org.apache.velocity", "1.7", "HIGH", "CVE-2020-13936", "Template injection via ClassInfo exposure", "Upgrade to velocity-engine-core:2.3"),
    ("velocity", "org.apache.velocity", "1.7", "HIGH", "CVE-2021-27807", "ReDoS via crafted regex in ClassInfo", "Upgrade to velocity-engine-core:2.3"),
    ("mysql-connector-j", "com.mysql", "managed", "MEDIUM", "CVE-2023-21971", "MITM with allowPublicKeyRetrieval=true", "Set allowPublicKeyRetrieval=false in JDBC URL"),
    ("spring-boot-starter-parent", "org.springframework.boot", "3.3.4", "INFO", "N/A", "Current stable — monitor for updates", "Upgrade to 3.4.x when stable"),
    ("axios", "npm", "^1.7.7", "LOW", "N/A", "No known critical CVEs", "Keep updated"),
    ("react", "npm", "^18.3.1", "LOW", "N/A", "No known critical CVEs", "Keep updated"),
    ("tailwindcss", "npm", "^3.4.13", "LOW", "N/A", "No known critical CVEs", "Keep updated"),
]

# ─────────────────────────────────────────────
# WORKBOOK 1: findings.xlsx
# ─────────────────────────────────────────────
def create_findings_workbook():
    wb = openpyxl.Workbook()

    # ── Sheet 1: Security Findings ──
    ws1 = wb.active
    ws1.title = "Security Findings"
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = "A3"

    # Title
    ws1.merge_cells("A1:L1")
    title_cell = ws1["A1"]
    title_cell.value = "🔒  SECURITY FINDINGS — Smart Parking & Reservation System"
    title_cell.fill = make_fill("1A1A2E")
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 36

    headers = ["ID", "Severity", "Type", "CWE", "OWASP", "File Path",
               "Endpoint", "Description", "Exploitation Scenario", "Impact", "Fix", "Status"]
    apply_header_row(ws1, 2, headers)
    ws1.row_dimensions[2].height = 28

    sev_colors = {"Critical": "C0392B", "High": "E67E22", "Medium": "F4D03F", "Low": "27AE60"}
    row_fill_alt = ["F9F9F9", "FFFFFF"]

    for i, f in enumerate(FINDINGS):
        row_num = i + 3
        sev = f["severity"]
        row_data = [
            f["id"], sev, f["type"], f["cwe"], f["owasp"],
            f["file"], f["endpoint"], f["description"],
            f["exploitation"], f["impact"], f["fix"], "OPEN"
        ]
        bg = row_fill_alt[i % 2]
        apply_data_row(ws1, row_num, row_data, fill_color=bg)

        # Severity cell
        sev_cell = ws1.cell(row=row_num, column=2)
        sev_cell.fill = make_fill(sev_colors.get(sev, "AAAAAA"))
        sev_cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        sev_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws1.row_dimensions[row_num].height = 60

    col_widths = [8, 12, 30, 12, 30, 40, 30, 50, 50, 30, 50, 12]
    set_col_widths(ws1, col_widths)

    # ── Sheet 2: Endpoint Inventory ──
    ws2 = wb.create_sheet("Endpoint Inventory")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A3"

    ws2.merge_cells("A1:G1")
    t = ws2["A1"]
    t.value = "🌐  API ENDPOINT INVENTORY — Smart Parking & Reservation"
    t.fill = make_fill("16213E")
    t.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 36

    ep_headers = ["HTTP Method", "Endpoint", "Auth Required", "Expected Roles", "Controller File", "Description", "Risk Level"]
    apply_header_row(ws2, 2, ep_headers, fill_color="16213E")
    ws2.row_dimensions[2].height = 28

    method_colors = {"GET": "2980B9", "POST": "27AE60", "PUT": "E67E22", "DELETE": "C0392B"}
    risk_by_auth = {"No ⚠️": ("C0392B", "HIGH RISK"), "No": ("27AE60", "Public"), }

    for i, ep in enumerate(ENDPOINTS):
        method, endpoint, auth, roles, ctrl, desc = ep
        row_num = i + 3
        risk = "HIGH RISK — No Auth" if "⚠️" in auth else "Public/OK"
        apply_data_row(ws2, row_num, [method, endpoint, auth, roles, ctrl, desc, risk],
                       fill_color=["F9F9F9", "FFFFFF"][i % 2])

        # Method cell color
        mc = ws2.cell(row=row_num, column=1)
        mc.fill = make_fill(method_colors.get(method, "555555"))
        mc.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        mc.alignment = Alignment(horizontal="center", vertical="center")

        # Risk cell color
        rc = ws2.cell(row=row_num, column=7)
        rc.fill = make_fill("C0392B" if "HIGH" in risk else "27AE60")
        rc.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        rc.alignment = Alignment(horizontal="center", vertical="center")

        ws2.row_dimensions[row_num].height = 30

    set_col_widths(ws2, [12, 40, 14, 20, 35, 50, 20])

    # ── Sheet 3: Dependency Vulnerabilities ──
    ws3 = wb.create_sheet("Dependency Vulnerabilities")
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = "A3"

    ws3.merge_cells("A1:G1")
    t3 = ws3["A1"]
    t3.value = "📦  DEPENDENCY VULNERABILITY REPORT"
    t3.fill = make_fill("0F3460")
    t3.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    t3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 36

    dep_headers = ["Package", "GroupId", "Version", "Severity", "CVE", "Description", "Recommended Fix"]
    apply_header_row(ws3, 2, dep_headers, fill_color="0F3460")
    ws3.row_dimensions[2].height = 28

    dep_sev_colors = {"HIGH": "C0392B", "MEDIUM": "E67E22", "LOW": "27AE60", "INFO": "2980B9"}

    for i, dep in enumerate(DEPENDENCIES):
        row_num = i + 3
        apply_data_row(ws3, row_num, list(dep), fill_color=["F9F9F9", "FFFFFF"][i % 2])
        sev_c = ws3.cell(row=row_num, column=4)
        sev_c.fill = make_fill(dep_sev_colors.get(dep[3], "888888"))
        sev_c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        sev_c.alignment = Alignment(horizontal="center", vertical="center")
        ws3.row_dimensions[row_num].height = 35

    set_col_widths(ws3, [25, 30, 15, 12, 18, 55, 45])

    # ── Sheet 4: Risk Summary ──
    ws4 = wb.create_sheet("Risk Summary")
    ws4.sheet_view.showGridLines = False

    ws4.merge_cells("A1:D1")
    t4 = ws4["A1"]
    t4.value = "📊  RISK SUMMARY DASHBOARD"
    t4.fill = make_fill("1A1A2E")
    t4.font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    t4.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 40

    # Summary stats
    summary_data = [
        ("", "", "", ""),
        ("METRIC", "VALUE", "COLOR CODE", "NOTES"),
        ("Overall Security Score", "18 / 100", "🔴 CRITICAL", "Immediate action required"),
        ("", "", "", ""),
        ("Critical Findings", "3", "🔴", "Wildcard CORS, No Auth, IDOR"),
        ("High Findings", "5", "🟠", "Path Traversal, Data Exposure, etc."),
        ("Medium Findings", "5", "🟡", "Headers, Debug mode, Upload limit"),
        ("Low Findings", "3", "🟢", "Stack traces, EOL dep, Status validation"),
        ("Total Findings", "16", "—", ""),
        ("", "", "", ""),
        ("OWASP A01 - Broken Access Control", "FAIL ❌", "🔴", "IDOR, No RBAC"),
        ("OWASP A02 - Cryptographic Failures", "FAIL ❌", "🔴", "HTTP, empty DB password"),
        ("OWASP A03 - Injection", "PARTIAL ⚠️", "🟠", "Path traversal risk"),
        ("OWASP A05 - Security Misconfiguration", "FAIL ❌", "🔴", "CORS wildcard, missing headers"),
        ("OWASP A06 - Vulnerable Components", "FAIL ❌", "🟠", "Velocity 1.7 EOL"),
        ("OWASP A07 - Auth Failures", "FAIL ❌", "🔴", "No server-side auth enforcement"),
        ("OWASP A09 - Security Logging", "FAIL ❌", "🟡", "e.printStackTrace(), show-sql=true"),
    ]

    for i, row in enumerate(summary_data, 2):
        for j, val in enumerate(row, 1):
            cell = ws4.cell(row=i, column=j, value=val)
            if i == 3:
                cell.fill = make_fill("1A1A2E")
                cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif i == 4:
                cell.fill = make_fill("2C3E50")
                cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif val:
                cell.fill = make_fill("F9F9F9" if i % 2 == 0 else "FFFFFF")
                cell.font = body_font()
                cell.alignment = Alignment(wrap_text=True)
            cell.border = make_border() if val else Border()
        ws4.row_dimensions[i].height = 22

    set_col_widths(ws4, [40, 20, 15, 45])

    return wb


# ─────────────────────────────────────────────
# WORKBOOK 2: endpoint-inventory.xlsx
# ─────────────────────────────────────────────
def create_endpoint_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "API Inventory"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "🌐  COMPLETE API ENDPOINT INVENTORY — Smart Parking & Reservation System"
    t.fill = make_fill("0D1117")
    t.font = Font(name="Calibri", bold=True, size=14, color="58A6FF")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    headers = ["#", "Method", "Endpoint", "Auth Required", "Expected Roles",
               "Controller", "Description", "Risk"]
    apply_header_row(ws, 2, headers, fill_color="161B22")
    ws.row_dimensions[2].height = 28

    method_colors = {"GET": "1E90FF", "POST": "28A745", "PUT": "FD7E14", "DELETE": "DC3545"}
    risk_colors   = {"HIGH RISK — No Auth": "C0392B", "Public/OK": "27AE60"}

    for i, ep in enumerate(ENDPOINTS):
        method, endpoint, auth, roles, ctrl, desc = ep
        risk = "HIGH RISK — No Auth" if "⚠️" in auth else "Public/OK"
        row_num = i + 3
        row_data = [i + 1, method, endpoint, auth.replace(" ⚠️", ""), roles, ctrl, desc, risk]
        apply_data_row(ws, row_num, row_data, fill_color=["F0F0F0", "FAFAFA"][i % 2])

        mc = ws.cell(row=row_num, column=2)
        mc.fill = make_fill(method_colors.get(method, "666666"))
        mc.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        mc.alignment = Alignment(horizontal="center", vertical="center")

        rc = ws.cell(row=row_num, column=8)
        rc.fill = make_fill(risk_colors.get(risk, "888888"))
        rc.font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
        rc.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row_num].height = 28

    set_col_widths(ws, [5, 10, 42, 14, 20, 35, 55, 22])
    return wb


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
if __name__ == "__main__":
    output_dir = os.path.join(os.path.dirname(__file__), "..", "Vulnerability Test Results")
    os.makedirs(output_dir, exist_ok=True)

    print("Generating findings.xlsx ...")
    wb1 = create_findings_workbook()
    wb1.save(os.path.join(output_dir, "findings.xlsx"))
    print("  ✅ findings.xlsx saved")

    print("Generating endpoint-inventory.xlsx ...")
    wb2 = create_endpoint_workbook()
    wb2.save(os.path.join(output_dir, "endpoint-inventory.xlsx"))
    print("  ✅ endpoint-inventory.xlsx saved")

    print("\nDone! Reports saved to 'Vulnerability Test Results/'")
